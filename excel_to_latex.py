#!/usr/bin/env python3
"""
Excel to LaTeX Table Converter
Converts Excel tables to LaTeX format with customizable styling
"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import re

# ============================================
# CONFIGURATION - Edit these values
# ============================================
INPUT_EXCEL_FILE = "complex.xlsx"  # Change this to your Excel file name
SHEET_NAME = None  # None = use active sheet, or specify sheet name like "Sheet1"
OUTPUT_LATEX_FILE = "output.tex"  # Output file name

# Table metadata (edit these or leave empty)
SECTION_TITLE = ""
TABLE_CAPTION = ""
TABLE_LABEL = ""
# ============================================


class ExcelToLatex:
    def __init__(self, excel_file, sheet_name=None):
        self.workbook = load_workbook(excel_file, data_only=True)
        if sheet_name:
            self.sheet = self.workbook[sheet_name]
        else:
            self.sheet = self.workbook.active
        self.merged_cells = self._get_merged_cells()
    
    def _get_merged_cells(self):
        """Get dictionary of merged cell ranges"""
        merged = {}
        for merge_range in self.sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = (
                merge_range.min_col, merge_range.min_row,
                merge_range.max_col, merge_range.max_row
            )
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged[(row, col)] = {
                        'min_row': min_row,
                        'min_col': min_col,
                        'max_row': max_row,
                        'max_col': max_col,
                        'rowspan': max_row - min_row + 1,
                        'colspan': max_col - min_col + 1
                    }
        return merged
    
    def _escape_latex(self, text):
        """Escape special LaTeX characters"""
        if text is None:
            return ""
        
        text = str(text)
        
        # Escape special characters
        replacements = {
            '\\': r'\textbackslash{}',
            '&': r'\&',
            '%': r'\%',
            '$': r'\$',
            '#': r'\#',
            '_': r'\_',
            '{': r'\{',
            '}': r'\}',
            '~': r'\textasciitilde{}',
            '^': r'\textasciicircum{}',
        }
        
        for old, new in replacements.items():
            text = text.replace(old, new)
        
        # Handle textbackslash that might have been double-escaped
        text = text.replace(r'\textbackslash{}\textbackslash{}', r'\textbackslash{}')
        
        return text
    
    def _detect_bullet_list(self, text):
        """Detect if cell content should be formatted as a bullet list"""
        if not text or not isinstance(text, str):
            return False, []
        
        # Check for Excel bullet character (•) or newlines with common list markers
        lines = text.split('\n')
        
        # Check if multiple lines start with bullet points
        bullet_chars = ['•', '·', '-', '*', '○', '▪']
        bullet_lines = []
        
        for line in lines:
            line = line.strip()
            if line:
                # Check if line starts with a bullet character
                starts_with_bullet = any(line.startswith(char) for char in bullet_chars)
                if starts_with_bullet:
                    # Remove the bullet character and any following spaces
                    for char in bullet_chars:
                        if line.startswith(char):
                            line = line[len(char):].strip()
                            break
                    bullet_lines.append(line)
                elif bullet_lines:
                    # If we already have bullets and this doesn't start with one,
                    # it might be a continuation of the previous item
                    bullet_lines[-1] += ' ' + line
        
        # If we found bullet points, return True
        if len(bullet_lines) >= 2:
            return True, bullet_lines
        
        # Check for numbered lists (1. 2. etc)
        numbered_pattern = re.compile(r'^\d+[\.\)]\s*(.+)$')
        numbered_lines = []
        for line in lines:
            line = line.strip()
            match = numbered_pattern.match(line)
            if match:
                numbered_lines.append(match.group(1))
        
        if len(numbered_lines) >= 2:
            return True, numbered_lines
        
        return False, []
    
    def _format_cell_content(self, text):
        """Format cell content, detecting and converting bullet lists"""
        if not text:
            return ""
        
        text = str(text)
        is_list, items = self._detect_bullet_list(text)
        
        if is_list:
            latex_items = [f"    \\item {self._escape_latex(item)}" for item in items]
            return "\n\\begin{itemize}\n" + "\n".join(latex_items) + "\n\\end{itemize}"
        else:
            # Replace newlines with LaTeX line breaks for regular text
            text = self._escape_latex(text)
            text = text.replace('\n', ' \\\\ ')
            return text
    
    def _get_cell_value(self, row, col):
        """Get cell value, handling merged cells"""
        cell_coord = (row, col)
        
        # If this is part of a merged cell but not the top-left
        if cell_coord in self.merged_cells:
            merge_info = self.merged_cells[cell_coord]
            if row == merge_info['min_row'] and col == merge_info['min_col']:
                # This is the top-left cell of merged range
                cell = self.sheet.cell(row, col)
                return self._format_cell_content(cell.value), merge_info
            else:
                # This is a continuation of merged cell, skip it
                return None, None
        
        cell = self.sheet.cell(row, col)
        return self._format_cell_content(cell.value), None
    
    def _calculate_column_widths(self, num_cols):
        """Calculate column widths based on Excel column widths, normalized to fit page"""
        # Get Excel column widths
        excel_widths = []
        for col_idx in range(1, num_cols + 1):
            col_letter = self.sheet.cell(1, col_idx).column_letter
            col_dim = self.sheet.column_dimensions[col_letter]
            # Excel default width is about 8.43, use that if not set
            width = col_dim.width if col_dim.width else 8.43
            excel_widths.append(width)
        
        # Calculate total width
        total_width = sum(excel_widths)
        
        # Maximum usable textwidth (leaving margins)
        MAX_TEXTWIDTH = 0.90
        
        # Normalize to fit within MAX_TEXTWIDTH
        normalized_widths = [(w / total_width) * MAX_TEXTWIDTH for w in excel_widths]
        
        # Round to 2 decimal places
        normalized_widths = [round(w, 2) for w in normalized_widths]
        
        # Ensure sum doesn't exceed MAX_TEXTWIDTH due to rounding
        current_sum = sum(normalized_widths)
        if current_sum > MAX_TEXTWIDTH:
            # Adjust the largest column slightly
            max_idx = normalized_widths.index(max(normalized_widths))
            normalized_widths[max_idx] -= (current_sum - MAX_TEXTWIDTH)
            normalized_widths[max_idx] = round(normalized_widths[max_idx], 2)
        
        return normalized_widths
    
    def generate_latex(self, section_title=None, caption=None, label=None):
        """Generate LaTeX table code"""
        # Get actual data range
        rows = list(self.sheet.iter_rows())
        if not rows:
            raise ValueError("Sheet is empty")
        
        # Find actual data bounds
        max_row = 0
        max_col = 0
        for i, row in enumerate(rows, 1):
            for j, cell in enumerate(row, 1):
                if cell.value is not None or isinstance(cell, MergedCell):
                    max_row = max(max_row, i)
                    max_col = max(max_col, j)
        
        if max_row == 0 or max_col == 0:
            raise ValueError("No data found in sheet")
        
        # Calculate column widths
        col_widths = self._calculate_column_widths(max_col)
        
        # Build column specification
        col_spec = "|" + "|".join([f">{{\\RaggedRight}}p{{{width}\\textwidth}}" 
                                     for width in col_widths]) + "|"
        
        # Start building LaTeX code
        latex = []
        latex.append("\\begin{table}[H]")
        
        if section_title:
            latex.append(f"\\section{{{section_title}}}")
        
        latex.append("\\centering")
        latex.append(f"\\begin{{tabular}}{{{col_spec}}}")
        latex.append("\\hline")
        
        # Process rows
        for row_idx in range(1, max_row + 1):
            row_data = []
            col_idx = 1
            
            while col_idx <= max_col:
                value, merge_info = self._get_cell_value(row_idx, col_idx)
                
                if value is None:
                    # Skip this cell (it's part of a merged cell)
                    col_idx += 1
                    continue
                
                if merge_info:
                    # Handle merged cell
                    colspan = merge_info['colspan']
                    if colspan > 1:
                        value = f"\\multicolumn{{{colspan}}}{{|c|}}{{{value}}}"
                    row_data.append(value)
                    col_idx += colspan
                else:
                    row_data.append(value)
                    col_idx += 1
            
            latex.append(" & ".join(row_data) + " \\\\ \\hline")
        
        latex.append("\\end{tabular}")
        
        if caption:
            latex.append(f"\\caption{{{caption}}}")
        
        if label:
            latex.append(f"\\label{{{label}}}")
        
        latex.append("\\end{table}")
        
        return "\n".join(latex)


def main():
    print("=" * 60)
    print("Excel to LaTeX Table Converter")
    print("=" * 60)
    
    # Validate input file
    input_path = Path(INPUT_EXCEL_FILE)
    if not input_path.exists():
        print(f"❌ Error: File '{INPUT_EXCEL_FILE}' not found")
        print(f"   Please edit INPUT_EXCEL_FILE in the script to point to your Excel file")
        return
    
    if input_path.suffix.lower() not in ['.xlsx', '.xls']:
        print("❌ Error: Input file must be .xlsx or .xls format")
        return
    
    try:
        print(f"📄 Reading Excel file: {INPUT_EXCEL_FILE}")
        if SHEET_NAME:
            print(f"📋 Using sheet: {SHEET_NAME}")
        else:
            print(f"📋 Using active sheet")
        
        # Convert
        converter = ExcelToLatex(INPUT_EXCEL_FILE, SHEET_NAME)
        latex_code = converter.generate_latex(
            section_title=SECTION_TITLE if SECTION_TITLE else None,
            caption=TABLE_CAPTION if TABLE_CAPTION else None,
            label=TABLE_LABEL if TABLE_LABEL else None
        )
        
        # Output
        output_path = Path(OUTPUT_LATEX_FILE)
        output_path.write_text(latex_code, encoding='utf-8')
        
        print(f"✅ LaTeX code generated successfully!")
        print(f"📝 Output written to: {OUTPUT_LATEX_FILE}")
        print()
        print("Preview:")
        print("-" * 60)
        # Show first 10 lines
        lines = latex_code.split('\n')
        for line in lines[:10]:
            print(line)
        if len(lines) > 10:
            print(f"... ({len(lines) - 10} more lines)")
        print("-" * 60)
    
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
